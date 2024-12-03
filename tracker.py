import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side

class TrackerSpreadsheetApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Tracker Spreadsheet Organizer")

        # Campaign code and file paths
        self.campaign_code_var = tk.StringVar()
        self.file_paths = {
            "Email 1 Opens": None,
            "Email 1 Clicks": None,
            "Email 2 Opens": None,
            "Email 2 Clicks": None,
            "Email 3 Opens": None,
            "Email 3 Clicks": None
        }

        # UI setup
        tk.Label(root, text="Campaign Code:").grid(row=0, column=0, padx=5, pady=5)
        tk.Entry(root, textvariable=self.campaign_code_var, width=30).grid(row=0, column=1, padx=5, pady=5)

        row_offset = 1
        for label in self.file_paths.keys():
            btn = tk.Button(root, text=f"Select {label} File", command=lambda l=label: self.select_file(l))
            btn.grid(row=row_offset, column=0, padx=5, pady=5, sticky="w")
            setattr(self, f"{label}_label", tk.Label(root, text="No file selected", fg="gray"))
            getattr(self, f"{label}_label").grid(row=row_offset, column=1, padx=5, pady=5, sticky="w")
            row_offset += 1

        # Process button
        self.process_btn = tk.Button(root, text="Process & Save Tracker Spreadsheet", command=self.process_files)
        self.process_btn.grid(row=row_offset, column=0, columnspan=2, padx=5, pady=10)

    def select_file(self, label):
        file_path = filedialog.askopenfilename(title=f"Select {label} CSV file", filetypes=[("CSV Files", "*.csv")])
        if file_path:
            self.file_paths[label] = file_path
            getattr(self, f"{label}_label").config(text=file_path.split('/')[-1], fg="black")

    def process_files(self):
        dataframes = {}

        # Collect only the files that are provided
        for label, file_path in self.file_paths.items():
            if file_path:  # Only process available files
                df = pd.read_csv(file_path)
                df.columns = df.columns.str.strip().str.lower()
                engagement_column = "opens" if label.endswith("Opens") else "clicks"
                required_columns = {
                    'email address': 'Email Address',
                    engagement_column: f'{label.replace(" ", "_")}',
                    'first name': 'First Name', 'last name': 'Last Name', 'address': 'Address',
                    'phone number': 'Phone Number', 'title': 'Title', 'company': 'Company',
                    'job position': 'Job Position', 'company website': 'Company Website',
                    'linkedin profile': 'LinkedIn Profile', 'member rating': 'Member Rating'
                }
                df = df.rename(columns=required_columns).reindex(columns=list(required_columns.values()), fill_value="")
                dataframes[label] = df

        # If no files are uploaded, create a blank spreadsheet
        if not dataframes:
            messagebox.showerror("No Input Files", "No files were provided. A blank spreadsheet will be created.")
            self.save_blank_spreadsheet()
            return

        # Consolidate all contacts and engagement data
        final_data = pd.concat(dataframes.values(), ignore_index=True)

        # Ensure all expected engagement columns are present
        engagement_columns = [
            'Email_1_Opens', 'Email_1_Clicks', 'Email_2_Opens', 'Email_2_Clicks', 'Email_3_Opens', 'Email_3_Clicks'
        ]
        contact_columns = [
            'Email Address', 'First Name', 'Last Name', 'Address', 'Phone Number', 'Title', 'Company',
            'Job Position', 'Company Website', 'LinkedIn Profile', 'Member Rating'
        ]

        # Add missing engagement columns and fill with 0
        for col in engagement_columns:
            if col not in final_data.columns:
                final_data[col] = 0

        # Group by "Email Address" and aggregate
        final_data = final_data.groupby('Email Address', as_index=False).agg({
            **{col: 'max' for col in engagement_columns},
            **{col: 'first' for col in contact_columns}
        })

        # Convert engagement columns to numeric, set non-numeric to 0
        for col in engagement_columns:
            final_data[col] = pd.to_numeric(final_data[col], errors='coerce').fillna(0)

        # Apply conditions to mark "Y" for Opens and Clicks
        for col in ['Email_1_Opens', 'Email_2_Opens', 'Email_3_Opens']:
            final_data[col] = final_data[col].apply(lambda x: 'Y' if x >= 2 else '')

        for col in ['Email_1_Clicks', 'Email_2_Clicks', 'Email_3_Clicks']:
            final_data[col] = final_data[col].apply(lambda x: 'Y' if x >= 1 else '')

        # Filter out rows with no engagement ('Y' values) in any engagement column
        final_data = final_data[final_data[engagement_columns].apply(lambda row: any(cell == 'Y' for cell in row), axis=1)]

        # Add additional columns and reorder as before
        final_data['Contact Status (for Jacqui to add)'] = ""
        final_data['Topic'] = "Engaged with Email Campaign"
        final_data['Status Reason'] = "New"
        final_data['Campaign Code'] = self.campaign_code_var.get()
        final_data['Lead Source'] = "Marketing Campaign"

        # Reorder columns as per requirements
        ordered_columns = [
            'Contact Status (for Jacqui to add)', 'Email_1_Opens', 'Email_1_Clicks', 'Email_2_Opens',
            'Email_2_Clicks', 'Email_3_Opens', 'Email_3_Clicks', 'Email Address', 'First Name', 'Last Name',
            'Address', 'Phone Number', 'Title', 'Company', 'Job Position', 'Company Website',
            'LinkedIn Profile', 'Member Rating', 'Topic', 'Status Reason', 'Campaign Code', 'Lead Source'
        ]
        final_data = final_data.reindex(columns=ordered_columns, fill_value="")

        # Key data with specified colors
        key_data = {
            "Key for Jacqui": ["sent follow up", "no response", "response"],
            "Colour": ["", "", ""]
        }
        key_df = pd.DataFrame(key_data)

        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if save_path:
            with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                key_df.to_excel(writer, sheet_name="Sheet1", index=False, startcol=0, startrow=0)
                final_data.to_excel(writer, sheet_name="Sheet1", index=False, startcol=6, startrow=0)

                workbook = writer.book
                sheet = workbook['Sheet1']

                # Bold headers in the key section
                sheet["A1"].font = Font(bold=True)
                sheet["B1"].font = Font(bold=True)

                # Apply colors to the key cells
                colors = {"yellow": "FFFF00", "red": "FF0000", "green": "00FF00"}
                color_rows = {"yellow": 2, "red": 3, "green": 4}
                for color, row in color_rows.items():
                    cell = sheet.cell(row=row, column=2)
                    cell.fill = PatternFill(start_color=colors[color], end_color=colors[color], fill_type="solid")

                # Add dividing borders between columns
                thin_border = Border(left=Side(style='thin'))
                for col in ["H", "J", "L", "N"]:
                    for row in range(2, len(final_data) + 2):
                        sheet[f"{col}{row}"].border = thin_border

            messagebox.showinfo("Success", f"Tracker Spreadsheet saved to {save_path}")

    def save_blank_spreadsheet(self):
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if save_path:
            with Workbook() as wb:
                wb.save(save_path)
            messagebox.showinfo("Success", f"Blank spreadsheet saved to {save_path}")


root = tk.Tk()
app = TrackerSpreadsheetApp(root)
root.mainloop()
